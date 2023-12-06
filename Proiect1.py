#pyautogui
import pyautogui
import time
import keyboard
import openpyxl
import pyperclip


def cautare_google():
    if pyautogui.locateOnScreen(r"C:\Users\catal\Pictures\Screenshots\Screenshotgoogle.png", confidence=0.2) !=None:
      print("da")
      camp_google=pyautogui.locateOnScreen(r"C:\Users\catal\Pictures\Screenshots\Screenshotgoogle.png", confidence=0.2)
      pyautogui.click(camp_google)
      time.sleep(2)
      pyautogui.write("https://translate.google.com/")
      pyautogui.press('enter')
      time.sleep(2)
    else:
       print ("IMAGINEA NU A FOST GASITA")
   

res = pyautogui.confirm("daca doriti sa incepeti rularea programului?","Confirmare")
if (res =="OK"):
    cautare_google()

    
def traducere_si_salvare_in_excel(cuvant):
    x = 157
    y = 256
    pyautogui.click(x,y)
    pyautogui.write(cuvant)
    time.sleep(2)
    x = 1074
    y = 427
    pyautogui.click(x,y)
    pyautogui.click(x,y)
    pyautogui.hotkey('ctrl','c')
    translated_world=pyperclip.paste()
    return translated_world

def salveaza_in_excel(cuvinte, nume_fisier="cuvinte_traduse.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Cuvinte traduse"
    sheet["A1"] = "Original"
    sheet["B2"] = "Traducere"

    row = 2
    for cuvant in cuvinte:
        sheet.cell(row, 1).value = cuvant
        sheet.cell(row, 2).value = traducere_si_salvare_in_excel(cuvant)
        row += 1
        
    workbook.save(nume_fisier)
    print("Traducerile au fost salvate în fișierul Excel.")

if pyautogui.confirm("Doriți să începeți rularea programului?", "Confirmare") == "OK":
    cautare_google()

    cuvinte_de_tradus = [
        "map\n\n,",
    
        # Adăugați mai multe cuvinte aici
    ]

    salveaza_in_excel(cuvinte_de_tradus)